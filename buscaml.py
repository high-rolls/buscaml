#! python3
# buscaml.py - Retorna resultados de busca no Mercado Livre
import datetime
import sys
import re
import requests
import bs4
import lxml
import openpyxl
import bitlyshortener

searchUrl = 'https://lista.mercadolivre.com.br/'
if len(sys.argv) > 1:
    searchUrl = searchUrl + '-'.join(sys.argv[1:])
else:
    print('Uso: buscaml.py [termos de busca]')
    exit(0)
print('Buscando produtos...')
res = requests.get(searchUrl)
res.raise_for_status()
wb = openpyxl.Workbook()
ws = wb.active
searchSoup = bs4.BeautifulSoup(res.text, 'lxml')
link_elems = searchSoup.select('a.ui-search-item__group__element.ui-search-link')
num_open = min(5, len(link_elems))
long_urls = []
# outFile.write('Codigo\tDescricao\tPreco\tQtd Vendida\tModelo\n'.encode('utf-8'))
#           A         B            C         D      E        F
ws.append(['Codigo', 'Descricao', 'Modelo', 'URL', 'Preco', 'Qtd Vendida'])
for i in range(num_open):
    print(f'Adicionando produto {i + 1} de {num_open}')
    lr = requests.get(link_elems[i].get('href'))
    lr.raise_for_status()
    row = str(i + 2)
    soup = bs4.BeautifulSoup(lr.text, 'lxml')
    id_match = re.search('MLB-?(\d+)', lr.url)
    if id_match:
        ws['A' + row] = int(id_match.group(1))
    title = soup.select('h1.ui-pdp-title')[0].get_text()
    ws['B' + row] = title
    model = ''
    model_th = soup.find('th', string='Modelo')
    if model_th and model_th.next_sibling:
        model = model_th.next_sibling.get_text()
    ws['C' + row] = model
    price_elem = soup.select('span.price-tag.ui-pdp-price__part .price-tag-amount')[0]
    price = price_elem.get_text()
    long_urls.append(lr.url)
    ws['E' + row] = price
    ws['E' + row].style = 'Currency [0]'
    # outFile.write(('\t' + price + '\n').encode('utf-8'))
    subtitle = soup.select('span.ui-pdp-subtitle')[0].getText()
    sold_match = re.search('\| +(\d+) +\w+', subtitle)
    sold_qty = 0
    if sold_match:
        sold_qty = sold_match.group(1)
    ws['F' + row] = int(sold_qty)
"""
print('Encurtando links...')
shortener = bitlyshortener.Shortener(tokens=[BITLY_TOKEN], max_cache_size=256)
short_urls = shortener.shorten_urls(long_urls)
for i in range(len(short_urls)):
    ws['D' + str(i + 2)] = short_urls[i]
"""
file_name = ' '.join(sys.argv[1:]) + ' ' + datetime.datetime.now().strftime('%Y%m%d %H%M%S') + '.xlsx'
print('Salvando arquivo ' + file_name)
wb.save('buscaml/' + file_name)
